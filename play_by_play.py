import pandas as pd
from openpyxl import Workbook

wb = Workbook()
ws = wb.active


pd.set_option('expand_frame_repr', False)

# Get data into a dataframe object
play_by_play_df = pd.read_csv('NBA Hackathon - Play by Play Data Sample (50 Games).txt', sep="\t")
event_codes_df = pd.read_csv('NBA Hackathon - Event Codes.txt', sep="\t")
game_lineup_df = pd.read_csv('NBA Hackathon - Game Lineup Data Sample (50 Games).txt', sep="\t")


# Sort the events in order specified in prompt
play_by_play_df.sort_values(['Period', 'PC_Time', 'WC_Time', 'Event_Num'], ascending=[True, False, True, True])

# Create map of event codes
#print(event_codes_df[:10])
game_ids = play_by_play_df.Game_id.unique()
periods = [1, 2, 3, 4]
games_list = []
count = 1
global_row = 2
ws['A1'] = 'Game_ID'
ws['B1'] = 'Player_ID'
ws['C1'] = 'Player_Plus/Minus'

# Repeat for every game	
for game_number in game_ids:
	
	print('GAME: ' + str(count))
	current_game = play_by_play_df.loc[play_by_play_df['Game_id'] == game_number]
	current_game_players = game_lineup_df.loc[game_lineup_df['Game_id'] == game_number]

	team_points = {}
	current_players_map = {}
	# Repeat for every period
	for period in periods:

		# current_period is dataframe of *play_by_play data* in current period
		current_period = current_game.loc[current_game['Period'] == period]

		# current_period_players is dataframe of *game_lineup data* in current period
		# reset index is needed because loc returns new df but with row # from original dataframe
		current_period_players = current_game_players.loc[current_game_players['Period'] == period].reset_index()

		# If not first period, means some players need to be reactivated/deactivated
		if period != 1:
			for players in current_players_map:
				current_players_map[players]['Active'] = False

		# Make players Active/status: Subbed on every period
		for i in range(len(current_period_players)):
			team = current_period_players['Team_id'][i]
			# Players dict: {'Player1': {'PM': +5, 'Active': True}}, if Active == False, means on the bench atm

			# If player not tracked yet (not subbed on yet/hasn't played), create their PM/Active values
			if current_period_players['Person_id'][i] not in current_players_map:
				current_players_map[current_period_players['Person_id'][i]] = {'PM': 0, 'Active': True, 'Team': team}
			# Otherwise if they are already tracked, if they're starting Active = True, else Active = False
			else:
				current_players_map[current_period_players['Person_id'][i]]['Active'] = True
			if team not in team_points.keys():
				team_points[team] = 0
	
		# Parse events ******************************************************************************************		
		event_msg_type = current_period['Event_Msg_Type']
		action_type = current_period['Action_Type']

		for index, events in current_period.iterrows():
			event_msg_type_description = event_codes_df.loc[(event_codes_df['Event_Msg_Type'] == event_msg_type[index]) & (event_codes_df['Action_Type'] == action_type[index])]['Event_Msg_Type_Description']
			# Turns series to list and get the description and not the index
			try:
				event_msg_type_description = list(event_msg_type_description)[0].strip()
			except:
				print('Event_Msg_Type invalid format!')
				print(events)
				input()
				continue

			if event_msg_type_description == 'Made Shot' or event_msg_type_description == 'Free Throw':
				# Print play by play data for every time there's a made shot
				team_points[events['Team_id']] += int(events['Option1'])
				for players in current_players_map:
					if current_players_map[players]['Active']:
						if events['Team_id'] == current_players_map[players]['Team']:
							current_players_map[players]['PM'] += int(events['Option1'])
						else:
							current_players_map[players]['PM'] -= int(events['Option1'])

			#if event_msg_type_description == 'Free Throw':


			if event_msg_type_description == 'Substitution':

				# ADD CLAUSE IN CASE SUBBING IN DURING FREE THROWS*******************
				current_players_map[events['Person1']]['Active'] = False
				if events['Person2'] not in current_players_map:
					current_players_map[events['Person2']] = {'PM': 0, 'Active': True, 'Team': events['Team_id']}
				else:
					current_players_map[events['Person2']]['Active'] = True	

	for i in current_players_map:
		ws['A'+str(global_row)] = str(game_number)
		ws['B'+str(global_row)] = str(i)
		ws['C'+str(global_row)] = str(current_players_map[i]['PM'])
		global_row += 1
		#print('Player ' + str(i) + ': ' + str(current_players_map[i]))
	print('Final Score: ' + str(team_points))
	count += 1

wb.save("OJSampson_Q1_BBALL.csv")